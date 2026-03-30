from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

wb = Workbook()

C_HEADER_BG   = "1F3864"
C_HEADER_FG   = "FFFFFF"
C_SECTION     = "2E75B6"
C_SECTION_FG  = "FFFFFF"
C_PASS        = "C6EFCE"
C_FAIL        = "FFC7CE"
C_BLOCKED     = "FFEB9C"
C_NOT_TESTED  = "F2F2F2"
C_ALT_ROW     = "EAF2FF"
C_BORDER      = "BFBFBF"

def fill(hex_colour):
    return PatternFill("solid", fgColor=hex_colour)

def bold(colour="000000", size=10):
    return Font(bold=True, color=colour, size=size)

def thin_border():
    s = Side(style="thin", color=C_BORDER)
    return Border(left=s, right=s, top=s, bottom=s)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

ws = wb.active
ws.title = "Test Cases"

HEADERS = [
    "Test Case ID", "User Story", "Role / Actor",
    "Description", "Pre-Conditions", "Test Steps",
    "Expected Result", "Actual Result", "Status",
    "Priority", "Notes / Defect Ref"
]

COL_WIDTHS = [14, 28, 16, 40, 34, 50, 40, 36, 14, 12, 30]

ws.freeze_panes = "A3"
ws.sheet_view.zoomScale = 90

ws.merge_cells("A1:K1")
title_cell = ws["A1"]
title_cell.value = "Andritz Vendor Portal — User Acceptance Test (UAT) Sheet"
title_cell.font  = Font(bold=True, color=C_HEADER_FG, size=14)
title_cell.fill  = fill(C_HEADER_BG)
title_cell.alignment = center()
ws.row_dimensions[1].height = 30

for col_idx, h in enumerate(HEADERS, start=1):
    cell = ws.cell(row=2, column=col_idx, value=h)
    cell.font      = bold(C_HEADER_FG, 10)
    cell.fill      = fill(C_HEADER_BG)
    cell.alignment = center()
    cell.border    = thin_border()
ws.row_dimensions[2].height = 32

for idx, w in enumerate(COL_WIDTHS, start=1):
    ws.column_dimensions[get_column_letter(idx)].width = w

SECTION = "SECTION"

data = [

  # ── 1. Authentication ─────────────────────────────────────────────────────
  (SECTION, "1. AUTHENTICATION"),

  ("TC-AUTH-01", "User Login", "All Roles",
   "Valid credentials login",
   "App is running; user account exists",
   "1. Open the app\n2. Enter valid email & password\n3. Click Login",
   "User logged in; JWT stored in localStorage; correct role dashboard shown",
   "", "Not Tested", "Critical", ""),

  ("TC-AUTH-02", "User Login", "All Roles",
   "Invalid password shows error",
   "App is running",
   "1. Enter valid email\n2. Enter wrong password\n3. Click Login",
   "Error message displayed; user NOT logged in",
   "", "Not Tested", "High", ""),

  ("TC-AUTH-03", "User Login", "All Roles",
   "Empty fields validation",
   "App is running",
   "1. Leave email and password blank\n2. Click Login",
   "Validation message shown; form not submitted",
   "", "Not Tested", "High", ""),

  ("TC-AUTH-04", "Session Management", "All Roles",
   "Expired / missing token redirects to login",
   "User is logged in",
   "1. Delete token from localStorage\n2. Refresh page",
   "User redirected to Login page; no data exposed",
   "", "Not Tested", "Critical", ""),

  ("TC-AUTH-05", "User Login", "All Roles",
   "Quick-fill demo buttons populate credentials",
   "Login page is open",
   "1. Click each demo account button\n2. Verify email & password fields fill correctly",
   "Correct credentials populated for each role",
   "", "Not Tested", "Medium", ""),

  # ── 2. Admin — User Management ────────────────────────────────────────────
  (SECTION, "2. ADMIN — USER MANAGEMENT"),

  ("TC-ADMIN-01", "Manage Users", "Admin",
   "Admin can view all active users",
   "Logged in as admin@andritz.com (Admin)",
   "1. Go to User Management tab\n2. Observe user table",
   "Table lists all active users with name, email, role, and designation columns",
   "", "Not Tested", "High", ""),

  ("TC-ADMIN-02", "Manage Users", "Admin",
   "Admin can create a new Buyer account",
   "Logged in as Admin",
   "1. Click 'Add User'\n2. Fill name, email (@andritz.com), password, role=Buyer\n3. Submit",
   "New user appears in the table; user can log in with new credentials",
   "", "Not Tested", "Critical", ""),

  ("TC-ADMIN-03", "Manage Users", "Admin",
   "Admin can create a new Approver account",
   "Logged in as Admin",
   "1. Click 'Add User'\n2. Fill details, role=Approver\n3. Submit",
   "New Approver in table; visible in Buyer's approver dropdown",
   "", "Not Tested", "Critical", ""),

  ("TC-ADMIN-04", "Manage Users", "Admin",
   "Duplicate email is rejected with Conflict error",
   "Logged in as Admin; existing user email known",
   "1. Click 'Add User'\n2. Enter an email that already exists\n3. Submit",
   "Error shown: 'email already in use'; no duplicate user created",
   "", "Not Tested", "High", ""),

  ("TC-ADMIN-05", "Manage Users", "Admin",
   "Non @andritz.com email is rejected",
   "Logged in as Admin",
   "1. Click 'Add User'\n2. Enter a gmail.com or other non-andritz email\n3. Submit",
   "Validation error: 'Email address must use the @andritz.com domain'",
   "", "Not Tested", "High", ""),

  ("TC-ADMIN-06", "Manage Users", "Admin",
   "Admin can edit an existing user's name, designation, and role",
   "Logged in as Admin; a Buyer account exists",
   "1. Click Edit on a user\n2. Change name, designation, role to Approver\n3. Save",
   "Changes saved; user's role updated; new role reflected in their next login",
   "", "Not Tested", "High", ""),

  ("TC-ADMIN-07", "Manage Users", "Admin",
   "Admin can reset a user's password",
   "Logged in as Admin; target user account exists",
   "1. Click Edit on a user\n2. Enter a new password in 'New Password' field\n3. Save",
   "Password updated; user can log in with new password",
   "", "Not Tested", "High", ""),

  ("TC-ADMIN-08", "Manage Users", "Admin",
   "Admin can archive (soft-delete) a user",
   "Logged in as Admin; a non-FA user exists",
   "1. Click Delete/Archive on a user\n2. Confirm",
   "User removed from active list; archived user does not appear in approver dropdown",
   "", "Not Tested", "High", ""),

  ("TC-ADMIN-09", "Manage Users", "Admin",
   "Admin can view archived users",
   "Logged in as Admin; at least one archived user exists",
   "1. Navigate to archived users view (if available)\n2. Observe list",
   "Archived users displayed separately from active users",
   "", "Not Tested", "Medium", ""),

  ("TC-ADMIN-10", "Manage Users", "Admin",
   "Pardeep Sharma (FinalApprover) cannot be edited or deleted via User Management",
   "Logged in as Admin",
   "1. Find pardeep.sharma@andritz.com in user list\n2. Attempt to edit or delete",
   "Error: 'The Final Approver account cannot be modified through User Management'",
   "", "Not Tested", "Critical", ""),

  ("TC-ADMIN-11", "Role Isolation", "Non-Admin",
   "Non-admin users cannot access User Management",
   "Logged in as Buyer or Approver",
   "1. Log in as Buyer\n2. Check that User Management is not visible or accessible",
   "User Management tab hidden; GET /api/users returns 403",
   "", "Not Tested", "Critical", ""),

  # ── 3. Admin — Request Oversight & Edits ─────────────────────────────────
  (SECTION, "3. ADMIN — REQUEST OVERSIGHT & EDITS"),

  ("TC-ADMIN-R01", "View All Requests", "Admin",
   "Admin sees all vendor requests across all statuses",
   "At least one request in each status exists",
   "1. Log in as Admin\n2. Open Requests tab\n3. Review list",
   "All requests visible; correct status badges shown for each",
   "", "Not Tested", "High", ""),

  ("TC-ADMIN-R02", "View All Requests", "Admin",
   "Stat cards (Total / Pending / Completed / Rejected) show correct counts",
   "Multiple requests in various statuses exist",
   "1. Open Admin Requests tab\n2. Check each stat card count",
   "Each card count matches the number of requests with that status in the table",
   "", "Not Tested", "Medium", ""),

  ("TC-ADMIN-R03", "Admin Edit Completed Request", "Admin",
   "Admin can edit vendor fields on a Completed request; revision entry is created",
   "At least one request in Completed status",
   "1. Open a Completed request\n2. Edit one or more fields (e.g. Contact Person)\n3. Save",
   "Fields updated; RevisionNo incremented; revision history shows [Admin] as changer with old→new values",
   "", "Not Tested", "High", ""),

  ("TC-ADMIN-R04", "Admin Edit Non-Completed", "Admin",
   "Admin cannot edit a request that is NOT in Completed status",
   "A request in PendingApproval or Rejected status exists",
   "1. Attempt to edit a non-Completed request as Admin",
   "Error: 'Only completed (SAP-approved) requests can be edited by Admin'",
   "", "Not Tested", "High", ""),

  ("TC-ADMIN-R05", "Classify Vendor", "Admin",
   "Admin can toggle vendor type between Permanent and One-Time on a Completed request",
   "A Completed request exists",
   "1. Open a Completed request\n2. Toggle the One-Time Vendor classification\n3. Save",
   "IsOneTimeVendor flag updated; classification badge changes in the UI",
   "", "Not Tested", "Medium", ""),

  # ── 4. Buyer — Create Vendor Request ─────────────────────────────────────
  (SECTION, "4. BUYER — CREATE VENDOR REQUEST"),

  ("TC-BUY-01", "Create Request", "Buyer",
   "Buyer can open new vendor registration form",
   "Logged in as vikram.nair@andritz.com (Buyer)",
   "1. Click 'New Vendor Request'\n2. Observe form",
   "Form opens with all required fields visible",
   "", "Not Tested", "Critical", ""),

  ("TC-BUY-02", "Create Request", "Buyer",
   "All mandatory fields validated before submit",
   "Vendor form is open",
   "1. Leave required fields blank\n2. Click Submit",
   "Validation errors shown per missing field; form not submitted",
   "", "Not Tested", "High", ""),

  ("TC-BUY-03", "Create Request", "Buyer",
   "City and Locality dropdowns load and filter correctly",
   "Vendor form is open",
   "1. Click City dropdown\n2. Select a city\n3. Check Locality dropdown",
   "City list populates; selecting a city filters Locality options accordingly",
   "", "Not Tested", "High", ""),

  ("TC-BUY-04", "Create Request", "Buyer",
   "Approver selection dropdown lists only Approver-role users",
   "Vendor form is open; at least one Approver account exists",
   "1. Click Approver selection\n2. Observe list",
   "Only Approver-role users shown; no Buyers, Admins, or FinalApprovers listed",
   "", "Not Tested", "Critical", ""),

  ("TC-BUY-05", "Create Request", "Buyer",
   "Successful submission creates request in PendingApproval status",
   "All form fields filled; approver selected",
   "1. Fill all fields\n2. Select approver(s)\n3. Submit",
   "Request created; status = PendingApproval; visible in Buyer's request list",
   "", "Not Tested", "Critical", ""),

  ("TC-BUY-06", "View Own Requests", "Buyer",
   "Buyer only sees their own requests (data isolation)",
   "Two buyer accounts each have requests",
   "1. Log in as Buyer A\n2. Check request list",
   "Only Buyer A's requests visible; other buyers' requests not shown",
   "", "Not Tested", "Critical", ""),

  ("TC-BUY-07", "View Request Detail", "Buyer",
   "Buyer can view full vendor detail, approval steps, and current status",
   "At least one request exists",
   "1. Click on a request\n2. Review detail modal/page",
   "All vendor fields, status, approval steps, and revision history visible",
   "", "Not Tested", "High", ""),

  # ── 5. Buyer — Rejection & Resubmission ──────────────────────────────────
  (SECTION, "5. BUYER — REJECTION & RESUBMISSION"),

  ("TC-REJ-01", "Rejection Flow", "Buyer",
   "Buyer sees rejection comments after Approver rejects",
   "Request is in Rejected status",
   "1. Log in as Buyer\n2. Open rejected request\n3. Check comments",
   "Rejection comment from Approver displayed on the request",
   "", "Not Tested", "Critical", ""),

  ("TC-REJ-02", "Resubmit Request", "Buyer",
   "Buyer can edit a rejected request and resubmit",
   "Request is in Rejected status",
   "1. Open rejected request\n2. Edit one or more fields\n3. Click Resubmit",
   "Request resubmitted; status = PendingApproval; RevisionNo incremented; all approval steps reset to Pending",
   "", "Not Tested", "Critical", ""),

  ("TC-REJ-03", "Revision History", "Buyer",
   "Revision history shows changed fields (old → new) after resubmit",
   "Request has been resubmitted at least once",
   "1. Open request with RevisionNo >= 1\n2. View Revision History section",
   "REV 1 entry shown; changed field names listed with old and new values; changed-by name and timestamp shown",
   "", "Not Tested", "High", ""),

  ("TC-REJ-04", "Resubmit on Non-Rejected", "Buyer",
   "Buyer cannot resubmit a request that is not in Rejected status",
   "A request in PendingApproval or Completed status",
   "1. Attempt to resubmit a non-Rejected request",
   "Error: 'Only Rejected requests can be resubmitted'; request unchanged",
   "", "Not Tested", "High", ""),

  # ── 6. Buyer — Update Completed Request ──────────────────────────────────
  (SECTION, "6. BUYER — UPDATE COMPLETED REQUEST"),

  ("TC-BUY-UPD-01", "Update Completed Request", "Buyer",
   "Buyer can update vendor fields on their own Completed request",
   "A request created by this Buyer is in Completed status",
   "1. Open a Completed request\n2. Edit one or more fields\n3. Save",
   "Fields updated; status remains Completed; vendor code preserved; RevisionNo incremented",
   "", "Not Tested", "High", ""),

  ("TC-BUY-UPD-02", "Update Completed — Revision Created", "Buyer",
   "Revision entry recorded when Buyer updates Completed request",
   "Buyer has just updated a Completed request",
   "1. Open the request\n2. View Revision History",
   "New revision entry visible with changed fields (old → new values) and timestamp",
   "", "Not Tested", "High", ""),

  ("TC-BUY-UPD-03", "Update Completed — Another Buyer Blocked", "Buyer",
   "Buyer cannot update another buyer's Completed request",
   "Two buyer accounts exist; Buyer B tries to update Buyer A's completed request",
   "1. Log in as Buyer B\n2. Attempt to update Buyer A's completed request",
   "403 Forbidden; request unchanged",
   "", "Not Tested", "Critical", ""),

  # ── 7. Approver — Review Workflow ─────────────────────────────────────────
  (SECTION, "7. APPROVER — REVIEW WORKFLOW"),

  ("TC-APP-01", "Pending Requests", "Approver",
   "Approver sees only requests assigned to them",
   "Logged in as rajesh.kumar@andritz.com (Approver); request assigned to them",
   "1. Log in as Approver\n2. Check Pending tab",
   "Only requests where this Approver is the current pending step are shown",
   "", "Not Tested", "Critical", ""),

  ("TC-APP-02", "Approve Request", "Approver",
   "Approver can approve a pending request",
   "Request in PendingApproval assigned to this Approver as the active step",
   "1. Open pending request\n2. Click Approve\n3. Confirm",
   "Request advances workflow; removed from Approver's Pending list; moves to History",
   "", "Not Tested", "Critical", ""),

  ("TC-APP-03", "Reject Request", "Approver",
   "Approver can reject a request with a mandatory comment",
   "Request in PendingApproval assigned to this Approver",
   "1. Open pending request\n2. Click Reject\n3. Enter rejection comment\n4. Confirm",
   "Status = Rejected; comment saved; request in Approver's History with Rejected badge",
   "", "Not Tested", "Critical", ""),

  ("TC-APP-04", "Reject Without Comment", "Approver",
   "Rejection requires a non-empty comment",
   "Request in PendingApproval",
   "1. Click Reject\n2. Leave comment blank\n3. Try to submit",
   "Validation error shown; rejection not processed without a comment",
   "", "Not Tested", "High", ""),

  ("TC-APP-05", "Approval History", "Approver",
   "History tab shows all requests the Approver has acted upon",
   "Approver has previously approved or rejected at least one request",
   "1. Go to History tab\n2. Review list",
   "Acted-upon requests listed with decision badge (Approved/Rejected) and comment",
   "", "Not Tested", "High", ""),

  ("TC-APP-06", "Sequential Blocking — Multi-Step", "Approver",
   "Second approver cannot act until first approver has approved",
   "Request assigned to Approver A (step 1) and Approver B (step 2); neither has acted",
   "1. Log in as Approver B\n2. Check Pending tab",
   "Request does NOT appear in Approver B's Pending list while Approver A's step is still Pending",
   "", "Not Tested", "Critical", ""),

  ("TC-APP-07", "Multi-Step — Status Stays Pending", "Approver",
   "Status stays PendingApproval after first of two approvers approves",
   "Request has two intermediate approvers; Approver A approves",
   "1. Log in as Approver A\n2. Approve the request\n3. Check request status",
   "Status remains PendingApproval (not yet PendingFinalApproval); Approver B's step now active",
   "", "Not Tested", "Critical", ""),

  ("TC-APP-08", "Multi-Step — Advances to Final", "Approver",
   "After ALL intermediate approvers approve, status moves to PendingFinalApproval",
   "Request has two approvers; Approver A has already approved",
   "1. Log in as Approver B\n2. Approve the request\n3. Check status",
   "Status = PendingFinalApproval; request appears in Final Approver's Pending tab",
   "", "Not Tested", "Critical", ""),

  ("TC-APP-09", "Role Isolation", "Non-Approver",
   "Non-Approver (e.g. Buyer) cannot call the approve endpoint",
   "Logged in as Buyer",
   "1. Attempt POST /api/vendor-requests/{id}/approve with Buyer's JWT",
   "403 Forbidden returned; request not approved",
   "", "Not Tested", "Critical", ""),

  # ── 8. Final Approver ─────────────────────────────────────────────────────
  (SECTION, "8. FINAL APPROVER — FINAL APPROVAL & VENDOR CODE"),

  ("TC-FA-01", "Access Control", "FinalApprover",
   "Only pardeep.sharma@andritz.com can access the FinalApprover console",
   "Logged in as pardeep.sharma@andritz.com",
   "1. Log in\n2. Observe dashboard",
   "FinalApprover console shown with Pending and History tabs",
   "", "Not Tested", "Critical", ""),

  ("TC-FA-02", "Email Gate", "FinalApprover",
   "A FinalApprover-role account with a different email is denied",
   "A second FinalApprover-role account with a non-Pardeep email exists (test only)",
   "1. Log in with that account\n2. Attempt to complete a request",
   "Console hidden or Complete endpoint returns 403",
   "", "Not Tested", "Critical", ""),

  ("TC-FA-03", "Pending Final Requests", "FinalApprover",
   "Pending tab shows requests in PendingFinalApproval status",
   "At least one request has passed all intermediate approvers",
   "1. Log in as Pardeep\n2. Check Pending tab",
   "Requests in PendingFinalApproval listed",
   "", "Not Tested", "Critical", ""),

  ("TC-FA-04", "Complete with Vendor Code", "FinalApprover",
   "Final Approver enters SAP vendor code and completes the request",
   "Request in PendingFinalApproval",
   "1. Open request\n2. Enter vendor code (e.g. V-10045)\n3. Click Complete",
   "Status = Completed; vendorCode stored on record; assigned-by name and timestamp set",
   "", "Not Tested", "Critical", ""),

  ("TC-FA-05", "Complete Without Vendor Code", "FinalApprover",
   "Vendor code field is mandatory — blank code blocked",
   "Request in PendingFinalApproval",
   "1. Open request\n2. Leave vendor code blank\n3. Click Complete",
   "Validation error; request not completed",
   "", "Not Tested", "High", ""),

  ("TC-FA-06", "Duplicate Vendor Code Rejected", "FinalApprover",
   "Cannot assign a vendor code already used on another Completed request",
   "At least one Completed request with a known vendor code exists",
   "1. Open a new PendingFinalApproval request\n2. Enter the same vendor code\n3. Click Complete",
   "Error: 'Vendor code is already assigned to another request'; request not completed",
   "", "Not Tested", "Critical", ""),

  ("TC-FA-07", "Completed Request Shows Vendor Code to Buyer", "Buyer",
   "Buyer can see the assigned vendor code after completion",
   "Request has been completed by Final Approver",
   "1. Log in as Buyer\n2. Open completed request",
   "Vendor code, assigned-by name, and date visible on the request detail",
   "", "Not Tested", "High", ""),

  ("TC-FA-08", "Final Approver History", "FinalApprover",
   "History tab lists requests the Final Approver has acted upon",
   "At least one request completed or rejected by Final Approver",
   "1. Log in as Pardeep\n2. Check History tab",
   "Completed/Rejected requests listed with vendor codes, decision, and dates",
   "", "Not Tested", "Medium", ""),

  # ── 9. Workflow Status & Audit Trail ─────────────────────────────────────
  (SECTION, "9. WORKFLOW STATUS & AUDIT TRAIL"),

  ("TC-WF-01", "Happy Path End-to-End", "All Roles",
   "Full workflow: Buyer creates → Approver approves → Final Approver completes",
   "Clean test data; fresh vendor name",
   "1. Buyer creates & submits request\n2. Approver approves\n3. Final Approver enters vendor code and completes\n4. Verify status at each step",
   "Status transitions: PendingApproval → PendingFinalApproval → Completed",
   "", "Not Tested", "Critical", ""),

  ("TC-WF-02", "Rejection & Resubmission Path", "All Roles",
   "Approver rejects → Buyer resubmits → workflow restarts",
   "Request in PendingApproval",
   "1. Approver rejects with comment\n2. Buyer edits and resubmits\n3. Check status and RevisionNo",
   "Status: Rejected → PendingApproval; RevisionNo incremented; all approval steps reset to Pending",
   "", "Not Tested", "Critical", ""),

  ("TC-WF-03", "Approval Steps Audit", "All Roles",
   "Each approval action is recorded with approver name, decision, comment, and timestamp",
   "Request has been approved or rejected at least once",
   "1. Open a request that has been actioned\n2. View approval steps section",
   "Each step shows: approver name, decision (Approved/Rejected), comment, timestamp",
   "", "Not Tested", "High", ""),

  ("TC-WF-04", "Revision History Audit", "Buyer / Admin",
   "Revision history records each change with field-level diff",
   "Request has RevisionNo >= 1",
   "1. Open request with revision history\n2. Expand revision entries",
   "Each revision shows: rev number, who changed, when, old value → new value per field",
   "", "Not Tested", "High", ""),

  # ── 10. Profile Management ────────────────────────────────────────────────
  (SECTION, "10. PROFILE MANAGEMENT"),

  ("TC-PROF-01", "Update Display Name", "All Roles",
   "Any authenticated user can update their own display name",
   "Logged in as any user",
   "1. Open profile/settings\n2. Change full name\n3. Save",
   "Display name updated; new name shown in the UI on next load",
   "", "Not Tested", "Medium", ""),

  ("TC-PROF-02", "Change Own Password", "All Roles",
   "User can change their password by providing current and new password",
   "Logged in as any user",
   "1. Open profile\n2. Enter current password and a new password\n3. Save\n4. Log out and log back in with new password",
   "Password changed; login succeeds with new password",
   "", "Not Tested", "Medium", ""),

  ("TC-PROF-03", "Wrong Current Password Rejected", "All Roles",
   "Password change fails if current password is incorrect",
   "Logged in as any user",
   "1. Open profile\n2. Enter wrong current password with a new password\n3. Save",
   "Error returned; password NOT changed",
   "", "Not Tested", "High", ""),

  # ── 11. Security & Access Control ────────────────────────────────────────
  (SECTION, "11. SECURITY & ACCESS CONTROL"),

  ("TC-SEC-01", "JWT Required", "Unauthenticated",
   "All API endpoints return 401 without a token",
   "App is running; user is NOT logged in",
   "1. Make GET /api/vendor-requests without Authorization header",
   "401 Unauthorized returned",
   "", "Not Tested", "Critical", ""),

  ("TC-SEC-02", "Role-Based Access — Buyer", "Buyer",
   "Buyer cannot call Approver-only endpoints",
   "Logged in as Buyer",
   "1. Attempt POST /api/vendor-requests/{id}/approve with Buyer's JWT",
   "403 Forbidden returned",
   "", "Not Tested", "Critical", ""),

  ("TC-SEC-03", "FinalApproverOnly Policy", "Approver",
   "Regular Approver cannot call the /complete endpoint",
   "Logged in as Approver",
   "1. Attempt POST /api/vendor-requests/{id}/complete with Approver JWT",
   "403 Forbidden returned",
   "", "Not Tested", "Critical", ""),

  ("TC-SEC-04", "Data Isolation", "Buyer",
   "Buyer cannot retrieve another buyer's requests via API",
   "Two buyer accounts with separate requests exist",
   "1. Log in as Buyer A\n2. GET /api/vendor-requests\n3. Verify only own requests returned",
   "Only Buyer A's requests in response; no cross-user data leakage",
   "", "Not Tested", "Critical", ""),

  ("TC-SEC-05", "FinalApprover Role Not Creatable via UI", "Admin",
   "Admin cannot create a FinalApprover role account through User Management",
   "Logged in as Admin",
   "1. Click 'Add User'\n2. Attempt to select FinalApprover as role",
   "FinalApprover is not an option in the role dropdown; API returns 400 if tried directly",
   "", "Not Tested", "Critical", ""),

  # ── 12. Pending Features (Future Testing) ─────────────────────────────────
  (SECTION, "12. PENDING FEATURES (Future Testing)"),

  ("TC-PDF-01", "PDF Generation", "Admin / FinalApprover",
   "Completed request can be downloaded as a PDF document",
   "Request in Completed status",
   "1. Open completed request\n2. Click 'Download PDF'",
   "PDF downloaded with all vendor details, approval records, and vendor code",
   "", "Not Tested", "High", "Not yet implemented — BRD Step 7"),

  ("TC-EMAIL-01", "Email Notification", "Buyer",
   "Buyer receives email after vendor code is assigned",
   "Request just completed with vendor code",
   "1. Complete a request as Final Approver\n2. Check Buyer's email inbox",
   "Email received: vendor approval confirmation + vendor code",
   "", "Not Tested", "High", "Not yet implemented — BRD Step 8"),

  ("TC-AD-01", "Azure AD Sync Placeholder", "Admin",
   "Sync from AD button triggers the placeholder endpoint without crashing",
   "Logged in as Admin",
   "1. Click 'Sync from Active Directory'\n2. Observe response",
   "Placeholder response returned; UI shows sync triggered; no unhandled error",
   "", "Not Tested", "Low", "Placeholder endpoint only — POST /api/users/sync-ad"),
]

# ── write rows ────────────────────────────────────────────────────────────
row = 3
alt = False
for item in data:
    if item[0] == SECTION:
        ws.merge_cells(f"A{row}:K{row}")
        sc = ws.cell(row=row, column=1, value=item[1])
        sc.font      = Font(bold=True, color=C_SECTION_FG, size=11)
        sc.fill      = fill(C_SECTION)
        sc.alignment = left()
        sc.border    = thin_border()
        ws.row_dimensions[row].height = 22
        row += 1
        alt = False
        continue

    tc_id, story, role, desc, pre, steps, exp, actual, status, priority, notes = item
    row_data = [tc_id, story, role, desc, pre, steps, exp, actual, status, priority, notes]
    row_fill = fill(C_ALT_ROW) if alt else fill("FFFFFF")

    for col_idx, val in enumerate(row_data, start=1):
        cell = ws.cell(row=row, column=col_idx, value=val)
        cell.border    = thin_border()
        cell.alignment = left() if col_idx not in (1, 3, 9, 10) else center()
        cell.fill      = row_fill

    status_cell = ws.cell(row=row, column=9)
    if status == "Pass":
        status_cell.fill = fill(C_PASS)
        status_cell.font = Font(bold=True, color="276221")
    elif status == "Fail":
        status_cell.fill = fill(C_FAIL)
        status_cell.font = Font(bold=True, color="9C0006")
    elif status == "Blocked":
        status_cell.fill = fill(C_BLOCKED)
        status_cell.font = Font(bold=True, color="7D4807")
    else:
        status_cell.fill = fill(C_NOT_TESTED)

    pri_cell = ws.cell(row=row, column=10)
    pri_cell.font = Font(bold=True, size=10)
    if priority == "Critical":
        pri_cell.font = Font(bold=True, color="9C0006", size=10)
    elif priority == "High":
        pri_cell.font = Font(bold=True, color="833C00", size=10)
    elif priority == "Medium":
        pri_cell.font = Font(bold=True, color="375623", size=10)

    ws.row_dimensions[row].height = 70
    row += 1
    alt = not alt

dv = DataValidation(
    type="list",
    formula1='"Not Tested,Pass,Fail,Blocked,N/A"',
    allow_blank=True,
    showDropDown=False
)
dv.error      = "Choose: Not Tested, Pass, Fail, Blocked, or N/A"
dv.errorTitle = "Invalid Status"
dv.prompt     = "Select test result"
dv.promptTitle= "Status"
dv.sqref      = f"I3:I{row}"
ws.add_data_validation(dv)

dvp = DataValidation(
    type="list",
    formula1='"Critical,High,Medium,Low"',
    allow_blank=True,
    showDropDown=False
)
dvp.sqref = f"J3:J{row}"
ws.add_data_validation(dvp)

# ════════════════════════════════════════════════════════════════════════════
# SHEET 2 – Summary Dashboard
# ════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Summary")
ws2.sheet_view.zoomScale = 110

ws2.merge_cells("A1:F1")
t = ws2["A1"]
t.value = "UAT Summary Dashboard"
t.font  = Font(bold=True, color=C_HEADER_FG, size=14)
t.fill  = fill(C_HEADER_BG)
t.alignment = center()
ws2.row_dimensions[1].height = 30

summary_headers = ["Area", "Total Cases", "Pass", "Fail", "Blocked", "Not Tested"]
for ci, h in enumerate(summary_headers, 1):
    c = ws2.cell(row=2, column=ci, value=h)
    c.font = bold(C_HEADER_FG, 10)
    c.fill = fill(C_SECTION)
    c.alignment = center()
    c.border = thin_border()
ws2.row_dimensions[2].height = 22

areas = [
    ("1. Authentication",                   5),
    ("2. Admin – User Management",          11),
    ("3. Admin – Request Oversight & Edits", 5),
    ("4. Buyer – Create Request",            7),
    ("5. Buyer – Rejection & Resubmission",  4),
    ("6. Buyer – Update Completed",          3),
    ("7. Approver – Review Workflow",        9),
    ("8. Final Approver",                    8),
    ("9. Workflow & Audit Trail",            4),
    ("10. Profile Management",               3),
    ("11. Security & Access Control",        5),
    ("12. Pending Features",                 3),
    ("TOTAL",                               67),
]

for ri, (area, total) in enumerate(areas, start=3):
    ws2.cell(row=ri, column=1, value=area).border    = thin_border()
    ws2.cell(row=ri, column=2, value=total).border   = thin_border()
    for ci in range(3, 7):
        c = ws2.cell(row=ri, column=ci, value=0 if area != "TOTAL" else "")
        c.border    = thin_border()
        c.alignment = center()
    if area == "TOTAL":
        ws2.cell(row=ri, column=1).font = Font(bold=True, size=10)
        ws2.cell(row=ri, column=2).font = Font(bold=True, size=10)
        ws2.cell(row=ri, column=1).fill = fill("D9E1F2")
        ws2.cell(row=ri, column=2).fill = fill("D9E1F2")
    else:
        ws2.cell(row=ri, column=1).alignment = left()
    ws2.row_dimensions[ri].height = 18

for ci, w in enumerate([32, 14, 10, 10, 12, 14], start=1):
    ws2.column_dimensions[get_column_letter(ci)].width = w

ws2.merge_cells("A17:F17")
ih = ws2["A17"]
ih.value = "HOW TO USE THIS SHEET"
ih.font  = Font(bold=True, color=C_HEADER_FG, size=11)
ih.fill  = fill(C_SECTION)
ih.alignment = center()
ws2.row_dimensions[17].height = 20

instructions = [
    ("1.", "Go to the 'Test Cases' tab and run each test case in order."),
    ("2.", "Follow the steps in column F 'Test Steps' exactly."),
    ("3.", "Record what actually happened in column H 'Actual Result'."),
    ("4.", "Set Status (column I) via dropdown: Pass / Fail / Blocked / Not Tested."),
    ("5.", "Add defect reference or observations in column K 'Notes'."),
    ("6.", "Update the Summary tab manually after completing each test area."),
    ("7.", "Prioritise Critical cases first, then High, then Medium / Low."),
]

for ri, (num, text) in enumerate(instructions, start=18):
    ws2.cell(row=ri, column=1, value=num).alignment   = center()
    ws2.cell(row=ri, column=1).border = thin_border()
    ws2.merge_cells(f"B{ri}:F{ri}")
    ws2.cell(row=ri, column=2, value=text).alignment  = left()
    ws2.cell(row=ri, column=2).border = thin_border()
    ws2.row_dimensions[ri].height = 18

ws2.merge_cells("A26:F26")
lh = ws2["A26"]
lh.value = "STATUS LEGEND"
lh.font  = Font(bold=True, color=C_HEADER_FG, size=11)
lh.fill  = fill(C_SECTION)
lh.alignment = center()
ws2.row_dimensions[26].height = 20

legend = [
    ("Pass",       C_PASS,       "276221", "Test executed and outcome matches expected result"),
    ("Fail",       C_FAIL,       "9C0006", "Test executed but outcome does NOT match expected result"),
    ("Blocked",    C_BLOCKED,    "7D4807", "Cannot execute — dependency or environment issue"),
    ("Not Tested", C_NOT_TESTED, "000000", "Test not yet executed"),
    ("N/A",        "FFFFFF",     "595959", "Not applicable in current scope"),
]
for ri, (status, bg, fg, meaning) in enumerate(legend, start=27):
    sc = ws2.cell(row=ri, column=1, value=status)
    sc.fill      = fill(bg)
    sc.font      = Font(bold=True, color=fg, size=10)
    sc.alignment = center()
    sc.border    = thin_border()
    ws2.merge_cells(f"B{ri}:F{ri}")
    mc = ws2.cell(row=ri, column=2, value=meaning)
    mc.alignment = left()
    mc.border    = thin_border()
    ws2.row_dimensions[ri].height = 18

# ════════════════════════════════════════════════════════════════════════════
# SHEET 3 – Test Accounts
# ════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Test Accounts")
ws3.sheet_view.zoomScale = 110

ws3.merge_cells("A1:E1")
ta = ws3["A1"]
ta.value = "Test Accounts — Andritz Vendor Portal"
ta.font  = Font(bold=True, color=C_HEADER_FG, size=14)
ta.fill  = fill(C_HEADER_BG)
ta.alignment = center()
ws3.row_dimensions[1].height = 30

acc_headers = ["Name", "Email", "Password", "Role", "Notes"]
for ci, h in enumerate(acc_headers, 1):
    c = ws3.cell(row=2, column=ci, value=h)
    c.font = bold(C_HEADER_FG, 10)
    c.fill = fill(C_SECTION)
    c.alignment = center()
    c.border = thin_border()
ws3.row_dimensions[2].height = 22

accounts = [
    ("Pardeep Sharma", "pardeep.sharma@andritz.com", "Change@Me1!",   "FinalApprover", "ONLY Final Approver — email-gated at role + policy level"),
    ("Vikram Nair",    "vikram.nair@andritz.com",    "Buyer@123!",    "Buyer",         "Default Buyer test account"),
    ("Rajesh Kumar",   "rajesh.kumar@andritz.com",   "Approver@123!", "Approver",      "Default Approver test account"),
    ("Admin",          "admin@andritz.com",           "Dahlia@1234",   "Admin",         "System Admin account (live DB credentials)"),
]
for ri, (name, email, pw, role, note) in enumerate(accounts, start=3):
    ws3.cell(row=ri, column=1, value=name).border  = thin_border()
    ws3.cell(row=ri, column=2, value=email).border = thin_border()
    ws3.cell(row=ri, column=3, value=pw).border    = thin_border()
    ws3.cell(row=ri, column=4, value=role).border  = thin_border()
    ws3.cell(row=ri, column=5, value=note).border  = thin_border()
    for ci in range(1, 6):
        ws3.cell(row=ri, column=ci).alignment = left()
    ws3.row_dimensions[ri].height = 20

for ci, w in enumerate([20, 34, 18, 16, 48], start=1):
    ws3.column_dimensions[get_column_letter(ci)].width = w

# ── save ──────────────────────────────────────────────────────────────────
OUT = r"c:\Users\16046\OneDrive\Desktop\Andritz BRD Dashboard\Andritz_Vendor_Portal_UAT_Test_Sheet.xlsx"
wb.save(OUT)
print(f"Saved: {OUT}")
